__version__ = '1.0.0'

import os
import pandas as pd
import numpy as np
import torch
import torch.nn as nn
import torch.optim as optim
from torch.utils.data import Dataset, DataLoader
from torch_geometric.nn import GCNConv
from scipy.spatial import KDTree
import openpyxl
from tqdm import tqdm
import math
import pickle
import random
from sklearn.cluster import KMeans
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
from sklearn.preprocessing import StandardScaler

# Environment setup
os.environ["OMP_NUM_THREADS"] = "1"
os.environ["OPENBLAS_NUM_THREADS"] = "1"
os.environ["MKL_NUM_THREADS"] = "1"
os.environ["NUMEXPR_NUM_THREADS"] = "1"
os.environ["VECLIB_MAXIMUM_THREADS"] = "1"

torch.set_num_threads(1)
torch.set_num_interop_threads(1)

device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
print(f'Using device: {device}')

def set_seed(seed=42):
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed(seed)
        torch.cuda.manual_seed_all(seed)
    torch.backends.cudnn.deterministic = True
    torch.backends.cudnn.benchmark = False

set_seed(42)

class GradientAwareLoss(nn.Module):
    def __init__(self, alpha=0.5, beta=0.3):
        super(GradientAwareLoss, self).__init__()
        self.alpha = alpha
        self.beta = beta
        self.mse = nn.MSELoss()
        
    def forward(self, pred, target):
        mse_loss = self.mse(pred, target)
        relative_error = torch.abs(pred - target) / (torch.abs(target) + 1e-6)
        rel_loss = torch.mean(relative_error)
        pred_grad = pred[1:] - pred[:-1]
        target_grad = target[1:] - target[:-1]
        grad_loss = self.mse(pred_grad, target_grad)
        return mse_loss + self.alpha * rel_loss + self.beta * grad_loss

class ST_TGN(nn.Module):
    def __init__(self, num_features, hidden_dim, num_layers, dropout=0.5, num_targets=1, num_heads=4):
        super(ST_TGN, self).__init__()
        self.gcn_layers = nn.ModuleList()
        self.gcn_layers.append(GCNConv(num_features, hidden_dim))
        self.dropout = nn.Dropout(dropout)
        for _ in range(num_layers - 1):
            self.gcn_layers.append(GCNConv(hidden_dim, hidden_dim))
        
        self.transformer_encoder_layer = nn.TransformerEncoderLayer(
            d_model=hidden_dim, 
            nhead=num_heads, 
            dropout=dropout,
            batch_first=True
        )
        self.transformer_encoder = nn.TransformerEncoder(
            self.transformer_encoder_layer, 
            num_layers=num_layers
        )
        
        self.fc = nn.Linear(hidden_dim, num_targets)
    
    def forward(self, x, edge_index, edge_attr, mask=None):
        batch_size, window_size, num_stations, num_features = x.shape
        x = x.view(-1, num_features)
        
        for conv in self.gcn_layers:
            x = conv(x, edge_index)
            x = torch.relu(x)
            x = self.dropout(x)
        
        x = x.view(batch_size, window_size, num_stations, -1)
        x = x.mean(dim=2)
        x = self.transformer_encoder(x, mask=mask)
        x = x[:, -1, :]
        out = self.fc(x)
        return out

class ClimateDataset(Dataset):
    def __init__(self, df, input_features, target, window_size=24, step_size=6, num_stations=1):
        super(ClimateDataset, self).__init__()
        self.target = target
        self.window_size = window_size
        self.step_size = step_size
        self.df = df
        self.input_features = input_features
        self.num_stations = num_stations
        self.windows = self.create_sliding_windows()
    
    def create_sliding_windows(self):
        data_list = []
        grouped = self.df.groupby('station_id')
        
        for station_id, group in grouped:
            group = group.sort_values('DATETIME').reset_index(drop=True)
            total_time_steps = len(group)
            
            for start in range(0, total_time_steps - self.window_size, self.step_size):
                end = start + self.window_size
                window = group.iloc[start:end]
                
                if len(window) != self.window_size:
                    continue
                
                x = window[self.input_features].values
                x = x.reshape(self.window_size, 1, -1)
                x = np.repeat(x, self.num_stations, axis=1)
                x = torch.tensor(x, dtype=torch.float)
                
                y = window[self.target].iloc[-1]
                y_tensor = torch.tensor([y], dtype=torch.float)
                
                data_list.append((x, y_tensor))
        
        return data_list
    
    def __len__(self):
        return len(self.windows)
    
    def __getitem__(self, idx):
        x, y = self.windows[idx]
        return x, y

def build_graph_kdtree(stations, K=5):
    coordinates = stations[['LATITUDE', 'LONGITUDE']].values
    kdtree = KDTree(coordinates)
    distances, indices = kdtree.query(coordinates, k=K)
    
    edge_index = []
    edge_attr = []
    
    for i in range(len(stations)):
        for j in range(1, K):
            neighbor = indices[i][j]
            distance = distances[i][j]
            edge_index.append([i, neighbor])
            edge_attr.append(distance)
    
    edge_index = torch.tensor(edge_index, dtype=torch.long).t().contiguous()
    edge_attr = torch.tensor(edge_attr, dtype=torch.float)
    
    return edge_index, edge_attr

def calculate_metrics(y_true, y_pred):
    bias = np.mean(y_pred - y_true)
    mae = mean_absolute_error(y_true, y_pred)
    rmse = np.sqrt(mean_squared_error(y_true, y_pred))
    r2 = r2_score(y_true, y_pred)
    return bias, mae, rmse, r2

def train_model(model, optimizer, criterion, train_loader, val_loader, num_epochs, target, excel_writer, sheet_name, save_dir, edge_index, edge_attr):
    best_val_loss = float('inf')
    best_val_r2 = -float('inf')
    no_improve_epochs_loss = 0
    no_improve_epochs_r2 = 0
    
    for epoch in range(num_epochs):
        model.train()
        total_loss = 0.0
        train_preds = []
        train_true = []
        
        for batch_x, batch_y in tqdm(train_loader, desc=f"Epoch {epoch+1}/{num_epochs} [Training]"):
            optimizer.zero_grad()
            x = batch_x.to(device)
            y = batch_y.to(device)
            output = model(x, edge_index, edge_attr)
            loss = criterion(output, y)
            loss.backward()
            optimizer.step()
            total_loss += loss.item()
            
            train_preds.append(output.detach().cpu().numpy())
            train_true.append(y.cpu().numpy())

        avg_train_loss = total_loss / len(train_loader)

        model.eval()
        val_loss = 0.0
        val_preds = []
        val_true = []
        
        with torch.no_grad():
            for batch_x, batch_y in tqdm(val_loader, desc=f"Epoch {epoch+1}/{num_epochs} [Validation]"):
                x = batch_x.to(device)
                y = batch_y.to(device)
                output = model(x, edge_index, edge_attr)
                loss = criterion(output, y)
                val_loss += loss.item()
                
                val_preds.append(output.detach().cpu().numpy())
                val_true.append(y.cpu().numpy())

        avg_val_loss = val_loss / len(val_loader)

        train_preds = np.concatenate(train_preds).flatten()
        train_true = np.concatenate(train_true).flatten()
        val_preds = np.concatenate(val_preds).flatten()
        val_true = np.concatenate(val_true).flatten()

        train_metrics = calculate_metrics(train_true, train_preds)
        val_metrics = calculate_metrics(val_true, val_preds)

        print(f'Epoch {epoch+1}/{num_epochs}, Train Loss: {avg_train_loss:.4f}, Val Loss: {avg_val_loss:.4f}')
        print(f'Train Metrics: Bias={train_metrics[0]:.4f}, MAE={train_metrics[1]:.4f}, RMSE={train_metrics[2]:.4f}, R2={train_metrics[3]:.4f}')
        print(f'Val Metrics: Bias={val_metrics[0]:.4f}, MAE={val_metrics[1]:.4f}, RMSE={val_metrics[2]:.4f}, R2={val_metrics[3]:.4f}')

        if avg_val_loss < best_val_loss:
            best_val_loss = avg_val_loss
            no_improve_epochs_loss = 0
            torch.save(model.state_dict(), os.path.join(save_dir, f'best_model_{target}.pth'))
        else:
            no_improve_epochs_loss += 1

        if val_metrics[3] > best_val_r2:
            best_val_r2 = val_metrics[3]
            no_improve_epochs_r2 = 0
        else:
            no_improve_epochs_r2 += 1

        if no_improve_epochs_loss >= 5:
            for param_group in optimizer.param_groups:
                param_group['lr'] *= 0.5
            print(f"Reducing learning rate to {optimizer.param_groups[0]['lr']}")
            no_improve_epochs_loss = 0

        if no_improve_epochs_r2 >= 10:
            print("Early stopping triggered")
            break

        if sheet_name not in excel_writer.sheetnames:
            sheet = excel_writer.create_sheet(sheet_name)
            headers = ['Epoch', 'Train Loss', 'Train Bias', 'Train MAE', 'Train RMSE', 'Train R2',
                       'Val Loss', 'Val Bias', 'Val MAE', 'Val RMSE', 'Val R2', 'Learning Rate']
            sheet.append(headers)
        else:
            sheet = excel_writer[sheet_name]

        row = [epoch+1, avg_train_loss] + list(train_metrics) + [avg_val_loss] + list(val_metrics) + [optimizer.param_groups[0]['lr']]
        sheet.append(row)

        excel_writer.save('path/training_results.xlsx')

    return best_val_loss, best_val_r2

__version__ = '1.0.0'

import os
import pandas as pd
import numpy as np
import torch
import torch.nn as nn
import torch.optim as optim
from torch.utils.data import Dataset, DataLoader
from torch_geometric.nn import GCNConv
from scipy.spatial import KDTree
import openpyxl
from tqdm import tqdm
import math
import pickle
import random
from sklearn.cluster import KMeans
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
from sklearn.preprocessing import StandardScaler

# Environment setup
os.environ["OMP_NUM_THREADS"] = "1"
os.environ["OPENBLAS_NUM_THREADS"] = "1"
os.environ["MKL_NUM_THREADS"] = "1"
os.environ["NUMEXPR_NUM_THREADS"] = "1"
os.environ["VECLIB_MAXIMUM_THREADS"] = "1"

torch.set_num_threads(1)
torch.set_num_interop_threads(1)

device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
print(f'Using device: {device}')

def set_seed(seed=42):
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed(seed)
        torch.cuda.manual_seed_all(seed)
    torch.backends.cudnn.deterministic = True
    torch.backends.cudnn.benchmark = False

set_seed(42)

class GradientAwareLoss(nn.Module):
    def __init__(self, alpha=0.5, beta=0.3):
        super(GradientAwareLoss, self).__init__()
        self.alpha = alpha
        self.beta = beta
        self.mse = nn.MSELoss()
        
    def forward(self, pred, target):
        mse_loss = self.mse(pred, target)
        relative_error = torch.abs(pred - target) / (torch.abs(target) + 1e-6)
        rel_loss = torch.mean(relative_error)
        pred_grad = pred[1:] - pred[:-1]
        target_grad = target[1:] - target[:-1]
        grad_loss = self.mse(pred_grad, target_grad)
        return mse_loss + self.alpha * rel_loss + self.beta * grad_loss

class ST_TGN(nn.Module):
    def __init__(self, num_features, hidden_dim, num_layers, dropout=0.5, num_targets=1, num_heads=4):
        super(ST_TGN, self).__init__()
        self.gcn_layers = nn.ModuleList()
        self.gcn_layers.append(GCNConv(num_features, hidden_dim))
        self.dropout = nn.Dropout(dropout)
        for _ in range(num_layers - 1):
            self.gcn_layers.append(GCNConv(hidden_dim, hidden_dim))
        
        self.transformer_encoder_layer = nn.TransformerEncoderLayer(
            d_model=hidden_dim, 
            nhead=num_heads, 
            dropout=dropout,
            batch_first=True
        )
        self.transformer_encoder = nn.TransformerEncoder(
            self.transformer_encoder_layer, 
            num_layers=num_layers
        )
        
        self.fc = nn.Linear(hidden_dim, num_targets)
    
    def forward(self, x, edge_index, edge_attr, mask=None):
        batch_size, window_size, num_stations, num_features = x.shape
        x = x.view(-1, num_features)
        
        for conv in self.gcn_layers:
            x = conv(x, edge_index)
            x = torch.relu(x)
            x = self.dropout(x)
        
        x = x.view(batch_size, window_size, num_stations, -1)
        x = x.mean(dim=2)
        x = self.transformer_encoder(x, mask=mask)
        x = x[:, -1, :]
        out = self.fc(x)
        return out

class ClimateDataset(Dataset):
    def __init__(self, df, input_features, target, window_size=24, step_size=6, num_stations=1):
        super(ClimateDataset, self).__init__()
        self.target = target
        self.window_size = window_size
        self.step_size = step_size
        self.df = df
        self.input_features = input_features
        self.num_stations = num_stations
        self.windows = self.create_sliding_windows()
    
    def create_sliding_windows(self):
        data_list = []
        grouped = self.df.groupby('station_id')
        
        for station_id, group in grouped:
            group = group.sort_values('DATETIME').reset_index(drop=True)
            total_time_steps = len(group)
            
            for start in range(0, total_time_steps - self.window_size, self.step_size):
                end = start + self.window_size
                window = group.iloc[start:end]
                
                if len(window) != self.window_size:
                    continue
                
                x = window[self.input_features].values
                x = x.reshape(self.window_size, 1, -1)
                x = np.repeat(x, self.num_stations, axis=1)
                x = torch.tensor(x, dtype=torch.float)
                
                y = window[self.target].iloc[-1]
                y_tensor = torch.tensor([y], dtype=torch.float)
                
                data_list.append((x, y_tensor))
        
        return data_list
    
    def __len__(self):
        return len(self.windows)
    
    def __getitem__(self, idx):
        x, y = self.windows[idx]
        return x, y

def build_graph_kdtree(stations, K=5):
    coordinates = stations[['LATITUDE', 'LONGITUDE']].values
    kdtree = KDTree(coordinates)
    distances, indices = kdtree.query(coordinates, k=K)
    
    edge_index = []
    edge_attr = []
    
    for i in range(len(stations)):
        for j in range(1, K):
            neighbor = indices[i][j]
            distance = distances[i][j]
            edge_index.append([i, neighbor])
            edge_attr.append(distance)
    
    edge_index = torch.tensor(edge_index, dtype=torch.long).t().contiguous()
    edge_attr = torch.tensor(edge_attr, dtype=torch.float)
    
    return edge_index, edge_attr

def calculate_metrics(y_true, y_pred):
    bias = np.mean(y_pred - y_true)
    mae = mean_absolute_error(y_true, y_pred)
    rmse = np.sqrt(mean_squared_error(y_true, y_pred))
    return bias, mae, rmse

def train_model(model, optimizer, criterion, train_loader, val_loader, num_epochs, target, excel_writer, sheet_name, save_dir, edge_index, edge_attr):
    best_val_loss = float('inf')
    no_improve_epochs = 0
    
    for epoch in range(num_epochs):
        model.train()
        total_loss = 0.0
        train_preds = []
        train_true = []
        
        for batch_x, batch_y in tqdm(train_loader, desc=f"Epoch {epoch+1}/{num_epochs} [Training]"):
            optimizer.zero_grad()
            x = batch_x.to(device)
            y = batch_y.to(device)
            output = model(x, edge_index, edge_attr)
            loss = criterion(output, y)
            loss.backward()
            optimizer.step()
            total_loss += loss.item()
            
            train_preds.append(output.detach().cpu().numpy())
            train_true.append(y.cpu().numpy())

        avg_train_loss = total_loss / len(train_loader)

        model.eval()
        val_loss = 0.0
        val_preds = []
        val_true = []
        
        with torch.no_grad():
            for batch_x, batch_y in tqdm(val_loader, desc=f"Epoch {epoch+1}/{num_epochs} [Validation]"):
                x = batch_x.to(device)
                y = batch_y.to(device)
                output = model(x, edge_index, edge_attr)
                loss = criterion(output, y)
                val_loss += loss.item()
                
                val_preds.append(output.detach().cpu().numpy())
                val_true.append(y.cpu().numpy())

        avg_val_loss = val_loss / len(val_loader)

        train_preds = np.concatenate(train_preds).flatten()
        train_true = np.concatenate(train_true).flatten()
        val_preds = np.concatenate(val_preds).flatten()
        val_true = np.concatenate(val_true).flatten()

        train_metrics = calculate_metrics(train_true, train_preds)
        val_metrics = calculate_metrics(val_true, val_preds)

        print(f'Epoch {epoch+1}/{num_epochs}, Train Loss: {avg_train_loss:.4f}, Val Loss: {avg_val_loss:.4f}')
        print(f'Train Metrics: Bias={train_metrics[0]:.4f}, MAE={train_metrics[1]:.4f}, RMSE={train_metrics[2]:.4f}')
        print(f'Val Metrics: Bias={val_metrics[0]:.4f}, MAE={val_metrics[1]:.4f}, RMSE={val_metrics[2]:.4f}')

        if avg_val_loss < best_val_loss:
            best_val_loss = avg_val_loss
            no_improve_epochs = 0
            torch.save(model.state_dict(), os.path.join(save_dir, f'best_model_{target}.pth'))
        else:
            no_improve_epochs += 1

        if no_improve_epochs >= 5:
            for param_group in optimizer.param_groups:
                param_group['lr'] *= 0.5
            print(f"Reducing learning rate to {optimizer.param_groups[0]['lr']}")
            no_improve_epochs = 0

        if no_improve_epochs >= 10:
            print("Early stopping triggered")
            break

        if sheet_name not in excel_writer.sheetnames:
            sheet = excel_writer.create_sheet(sheet_name)
            headers = ['Epoch', 'Train Loss', 'Train Bias', 'Train MAE', 'Train RMSE',
                       'Val Loss', 'Val Bias', 'Val MAE', 'Val RMSE', 'Learning Rate']
            sheet.append(headers)
        else:
            sheet = excel_writer[sheet_name]

        row = [epoch+1, avg_train_loss, train_metrics[0], train_metrics[1], train_metrics[2],
               avg_val_loss, val_metrics[0], val_metrics[1], val_metrics[2], optimizer.param_groups[0]['lr']]
        sheet.append(row)

        excel_writer.save('path/training_results.xlsx')

    return best_val_loss, best_val_r2

def main():
    # Load data
    df = pd.read_csv('path/Data.csv', parse_dates=['DATETIME'])
    
    if 'station_id' not in df.columns:
        stations = df[['LATITUDE', 'LONGITUDE']].drop_duplicates().reset_index(drop=True)
        stations['station_id'] = range(len(stations))
        df = df.merge(stations, on=['LATITUDE', 'LONGITUDE'], how='left')
    else:
        stations = df[['station_id', 'LATITUDE', 'LONGITUDE']].drop_duplicates().reset_index(drop=True)
    
    print(f"Number of unique stations: {len(stations)}")
    
    # Split stations using K-Means
    train_stations, val_stations, test_stations = split_stations(stations)
    
    train_val_df = df[df['station_id'].isin(train_stations + val_stations)].copy()
    test_df = df[df['station_id'].isin(test_stations)].copy()
    
    print(f"Training + Validation set size: {len(train_val_df)}")
    print(f"Test set size: {len(test_df)}")

    edge_index, edge_attr = build_graph_kdtree(stations, K=5)
    edge_index = edge_index.to(device)
    edge_attr = edge_attr.to(device)
    
    features = ['WRF_Wind_Speed_ms']
    target = 'Wind_Speed_ms'
    
    train_val_df = train_val_df.sort_values(['station_id', 'DATETIME'])
    test_df = test_df.sort_values(['station_id', 'DATETIME'])
    
    for df in [train_val_df, test_df]:
        df['WRF_Wind_Speed_ms_diff'] = df.groupby('station_id')['WRF_Wind_Speed_ms'].diff()
        df[f'{target}_diff'] = df.groupby('station_id')[target].diff()
    
    train_val_df = train_val_df.dropna()
    test_df = test_df.dropna()
    
    scaler = StandardScaler()
    train_val_df[features + [target]] = scaler.fit_transform(train_val_df[features + [target]])
    test_df[features + [target]] = scaler.transform(test_df[features + [target]])
    
    train_val_dataset = ClimateDataset(train_val_df, features + ['WRF_Wind_Speed_ms_diff'], target, num_stations=len(stations))
    train_size = int(0.8 * len(train_val_dataset))
    val_size = len(train_val_dataset) - train_size
    train_dataset, val_dataset = torch.utils.data.random_split(train_val_dataset, [train_size, val_size])
    
    train_loader = DataLoader(train_dataset, batch_size=32, shuffle=True, num_workers=0)
    val_loader = DataLoader(val_dataset, batch_size=32, shuffle=False, num_workers=0)
    
    # Initialize model
    model = ST_TGN(
        num_features=len(features) + 1,  # +1 for diff feature
        hidden_dim=64,
        num_layers=3,
        dropout=0.1,
        num_targets=1,
        num_heads=4
    ).to(device)
    
    # Define optimizer and loss
    optimizer = optim.Adam(model.parameters(), lr=0.001, weight_decay=1e-5)
    criterion = GradientAwareLoss()
    
    # Create Excel writer
    excel_writer = openpyxl.Workbook()
    
    # Create model save directory
    save_dir = 'path/models'
    os.makedirs(save_dir, exist_ok=True)
    
    # Train model
    num_epochs = 100
    best_val_loss = train_model(
        model=model,
        optimizer=optimizer,
        criterion=criterion,
        train_loader=train_loader,
        val_loader=val_loader,
        num_epochs=num_epochs,
        target=target,
        excel_writer=excel_writer,
        sheet_name=target,
        save_dir=save_dir,
        edge_index=edge_index,
        edge_attr=edge_attr
    )
    
    print(f'Best validation loss: {best_val_loss:.4f}')
    
    # Test model performance
    model.load_state_dict(torch.load(os.path.join(save_dir, f'best_model_{target}.pth')))
    model.eval()
    
    test_dataset = ClimateDataset(test_df, features + ['WRF_Wind_Speed_ms_diff'], target, num_stations=len(stations))
    test_loader = DataLoader(test_dataset, batch_size=32, shuffle=False, num_workers=0)
    
    test_preds = []
    test_true = []
    
    with torch.no_grad():
        for batch_x, batch_y in tqdm(test_loader, desc="Testing"):
            x = batch_x.to(device)
            y = batch_y.to(device)
            output = model(x, edge_index, edge_attr)
            test_preds.append(output.cpu().numpy())
            test_true.append(y.cpu().numpy())
    
    test_preds = np.concatenate(test_preds).flatten()
    test_true = np.concatenate(test_true).flatten()
    
    # Inverse transform predictions
    test_preds = scaler.inverse_transform(test_preds.reshape(-1, 1)).flatten()
    test_true = scaler.inverse_transform(test_true.reshape(-1, 1)).flatten()
    
    # Calculate test metrics
    test_bias, test_mae, test_rmse = calculate_metrics(test_true, test_preds)
    
    print(f'Test metrics for {target}:')
    print(f'  Bias: {test_bias:.4f}')
    print(f'  MAE: {test_mae:.4f}')
    print(f'  RMSE: {test_rmse:.4f}')
    
    # Save test results
    test_results = pd.DataFrame({
        'True_Values': test_true,
        'Predicted_Values': test_preds
    })
    test_results.to_csv('path/test_results.csv', index=False)
    
    # Save metrics to Excel
    test_sheet = excel_writer.create_sheet('Test Results')
    test_sheet.append(['Metric', 'Value'])
    test_sheet.append(['Bias', test_bias])
    test_sheet.append(['MAE', test_mae])
    test_sheet.append(['RMSE', test_rmse])
    
    # Calculate gradient metrics
    true_grad = np.diff(test_true)
    pred_grad = np.diff(test_preds)
    grad_mae = mean_absolute_error(true_grad, pred_grad)
    grad_rmse = np.sqrt(mean_squared_error(true_grad, pred_grad))
    
    # Add gradient metrics
    test_sheet.append(['Gradient MAE', grad_mae])
    test_sheet.append(['Gradient RMSE', grad_rmse])
    
    print(f'Gradient metrics:')
    print(f'  MAE: {grad_mae:.4f}')
    print(f'  RMSE: {grad_rmse:.4f}')
    
    excel_writer.save('path/wind_speed_results.xlsx')
    print("Results saved to wind_speed_results.xlsx")

if __name__ == '__main__':
    main()